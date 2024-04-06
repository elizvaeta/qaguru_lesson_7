import csv
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook
from paths import ZIP_FILE
from pypdf import PdfReader


def test_read_csv():
    with ZipFile(ZIP_FILE, 'r') as zip_file:
        with zip_file.open('test.csv') as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))

            assert csvreader[0][0] == 'Hello world!'


def test_read_xlsx():
    with ZipFile(ZIP_FILE, 'r') as zip_file:
        with zip_file.open('test.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=1, column=1).value == 'Hello world!'


def test_read_pdf():
    with ZipFile(ZIP_FILE, 'r') as zip_file:
        with zip_file.open('test.pdf') as pdf_file:
            reader = PdfReader(pdf_file)

            assert reader.pages[0].extract_text() == 'Hello world!'
